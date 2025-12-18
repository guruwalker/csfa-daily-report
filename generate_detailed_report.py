"""
Refactored CSFA Report Generator
Generates detailed Excel reports with visits, orders, and product details.
"""

import pandas as pd
import os
import logging
from typing import List, Dict, Any, Set, Tuple
from dataclasses import dataclass
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# Optional: dataframe_image for summary export
try:
    import dataframe_image as dfi
    HAS_DFI = True
except ImportError:
    HAS_DFI = False
    logging.warning("dataframe_image not installed. Summary image export disabled.")

from api_client import get_order_details

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


# ============================================================================
# CONFIGURATION
# ============================================================================

@dataclass
class ReportConfig:
    """Configuration for report generation."""
    output_file: str = "Daily_CSFA_Report.xlsx"
    summary_text_file: str = "summary_for_email.txt"
    summary_image_file: str = "summary_sheet.png"

    # Styling colors
    header_color: str = "4F81BD"
    customer_fill_color: str = "4BACC6"  # Bright blue
    product_header_color: str = "92D050"  # Light green
    error_color: str = "FF0000"

    # Fonts
    header_font_name: str = "Times New Roman"
    header_font_size: int = 12
    body_font_name: str = "Times New Roman"
    body_font_size: int = 12

    @classmethod
    def from_env(cls):
        """Create config from environment variables."""
        return cls(
            output_file=os.getenv("OUTPUT_FILE", "Daily_CSFA_Report.xlsx"),
            summary_text_file=os.getenv("SUMMARY_TEXT_FILE", "summary_for_email.txt"),
            summary_image_file=os.getenv("SUMMARY_IMAGE_FILE", "summary_sheet.png"),
        )


# ============================================================================
# DATA PROCESSING
# ============================================================================

class DataProcessor:
    """Handles data cleaning and processing."""

    @staticmethod
    def clean_visits(visits_data: List[Dict]) -> pd.DataFrame:
        """Clean and structure visits data."""
        visit_rows = []
        for v in visits_data:
            erp_code = v.get("erp_code") or ""
            visit_rows.append({
                "sales_rep": v.get("rep_name", ""),
                "customer_name": (v.get("shop_name") or "").strip(),
                "erp_code": erp_code.strip() if erp_code else "",
                "time_spent": v.get("timespent", "")
            })
        return pd.DataFrame(visit_rows)

    @staticmethod
    def clean_orders(orders_data: List[Dict]) -> pd.DataFrame:
        """Clean and structure orders data."""
        order_rows = []
        for o in orders_data:
            balance_str = o.get("balance", "0").replace(",", "").strip()
            try:
                balance = float(balance_str)
            except (ValueError, TypeError):
                balance = 0.0

            order_rows.append({
                "sales_rep": o.get("sales_rep", ""),
                "customer_name": (o.get("customer_name") or "").strip(),
                "customer_code": o.get("customer_code", ""),
                "order_id": o.get("id"),
                "order_value": balance
            })
        return pd.DataFrame(order_rows)

    @staticmethod
    def merge_visits_orders(
        df_visits: pd.DataFrame,
        df_orders: pd.DataFrame
    ) -> pd.DataFrame:
        """Merge visits and orders data with fallback logic."""
        # Merge by ERP code
        df_merge_code = pd.merge(
            df_visits,
            df_orders,
            left_on="erp_code",
            right_on="customer_code",
            how="left",
            suffixes=("_visit", "_order")
        )

        # Merge by customer name
        df_merge_name = pd.merge(
            df_visits,
            df_orders,
            left_on="customer_name",
            right_on="customer_name",
            how="left",
            suffixes=("_visit", "_order")
        )

        # Combine both merges
        df_final = df_merge_code.copy()
        for col in ["order_value", "customer_code", "sales_rep_order", "order_id"]:
            if col in df_merge_name.columns:
                df_final[col] = df_final[col].combine_first(df_merge_name[col])

        # Create final columns
        df_final["customer_name_final"] = df_final.get("customer_name_visit", pd.Series("")).combine_first(
            df_final.get("customer_name_order", pd.Series(""))
        )
        df_final["sales_rep_final"] = df_final.get("sales_rep_visit", pd.Series("")).combine_first(
            df_final.get("sales_rep_order", pd.Series(""))
        )

        return df_final

    @staticmethod
    def get_called_customers(
        df_visits: pd.DataFrame,
        df_orders: pd.DataFrame
    ) -> pd.DataFrame:
        """Find customers who were called but not visited."""
        visited_customers = set(df_visits["customer_name"])
        df_called = df_orders[~df_orders["customer_name"].isin(visited_customers)].copy()
        df_called["customer_called"] = df_called["customer_name"]
        return df_called

    @staticmethod
    def get_sales_reps(
        df_final: pd.DataFrame,
        df_called: pd.DataFrame
    ) -> List[str]:
        """Get sorted list of all sales representatives."""
        reps = set(df_final["sales_rep_final"].dropna()).union(
            df_called["sales_rep"].dropna()
        )
        return sorted(reps)


# ============================================================================
# EXCEL STYLING
# ============================================================================

class ExcelStyler:
    """Handles Excel worksheet styling."""

    def __init__(self, config: ReportConfig):
        self.config = config

    def calculate_row_height(self, text: str, column_width: float, font_size: int = 12) -> float:
        """
        Calculate the required row height for wrapped text.
        More accurate calculation based on actual character counting.

        Args:
            text: The text content
            column_width: Width of the column in Excel units
            font_size: Font size in points

        Returns:
            Required row height in points
        """
        if not text or pd.isna(text):
            return 15  # Default row height

        text = str(text)

        # Excel column width is measured in character widths
        # For Times New Roman 12pt, roughly 7 pixels per character
        # Column width in Excel is based on default font character width
        # Using 0.85 for more conservative estimate (ensures text fits)
        chars_per_line = max(1, int(column_width * 0.85))

        # Split by newlines first (in case there are explicit line breaks)
        lines = text.split('\n')
        total_lines = 0

        for line in lines:
            if len(line) == 0:
                total_lines += 1
            else:
                # Calculate how many wrapped lines this line will take
                line_count = max(1, int(len(line) / chars_per_line) + (1 if len(line) % chars_per_line > 0 else 0))
                total_lines += line_count

        # Calculate row height with more generous spacing
        # Base line height: font_size * 1.3 (generous line spacing)
        # Add padding: +8 points for readability
        line_height = font_size * 1.3
        calculated_height = (total_lines * line_height) + 8

        # Ensure minimum height of 18 points, maximum of 409 (Excel limit)
        return max(18, min(409, calculated_height))

    def apply_summary_styling(self, ws: Worksheet) -> None:
        """Apply styling to summary sheet (wrap text, wider columns, with borders)."""
        if not ws or ws.max_row == 0:
            logger.warning("Empty worksheet, skipping styling")
            return

        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Style header row
        header_font = Font(
            name=self.config.header_font_name,
            size=self.config.header_font_size,
            bold=True,
            color="FFFFFF"
        )
        header_fill = PatternFill(
            start_color=self.config.header_color,
            end_color=self.config.header_color,
            fill_type="solid"
        )
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        try:
            for col_idx, cell in enumerate(ws[1], start=1):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
                cell.border = thin_border
        except IndexError:
            logger.warning("Cannot style header row - worksheet may be empty")
            return

        # Style body rows
        body_font = Font(
            name=self.config.body_font_name,
            size=self.config.body_font_size
        )
        body_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = body_font
                cell.alignment = body_align
                cell.border = thin_border

        # Set wider column widths to minimize wrapping
        column_widths = {
            1: 25,  # SALESPERSON
            2: 20,  # CUSTOMERS VISITED
            3: 30,  # ORDER VALUE FROM VISITS
            4: 20,  # CUSTOMERS CALLED
            5: 30,  # ORDER VALUE FROM CALLS
        }

        for col_idx, width in column_widths.items():
            if col_idx <= ws.max_column:
                column = get_column_letter(col_idx)
                ws.column_dimensions[column].width = width

        # Set header row height (row 1)
        ws.row_dimensions[1].height = 30  # Fixed height for header

        # Calculate and set row heights for data rows (row 2 onwards)
        for row_idx in range(2, ws.max_row + 1):
            max_height = 18  # Default minimum height for data rows

            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value

                # Skip empty cells
                if cell_value is None or (isinstance(cell_value, str) and not cell_value.strip()):
                    continue

                # Get column width for this column
                col_width = column_widths.get(col_idx, 20)

                # Calculate required height
                try:
                    height = self.calculate_row_height(cell_value, col_width, self.config.body_font_size)
                    max_height = max(max_height, height)
                except Exception as e:
                    logger.warning(f"Could not calculate height for row {row_idx}, col {col_idx}: {e}")
                    max_height = max(max_height, 35)  # Fallback to larger height

            # Set the row height
            ws.row_dimensions[row_idx].height = max_height

    def _auto_adjust_columns(self, ws: Worksheet) -> None:
        """Auto-adjust column widths based on content."""
        for col_idx, col in enumerate(ws.columns, start=1):
            max_length = 0
            column = get_column_letter(col_idx)

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column].width = adjusted_width

    def format_money_columns(self, ws: Worksheet, columns: List[int]) -> None:
        """Format specific columns as money."""
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for col_idx in columns:
                if col_idx < len(row):
                    row[col_idx].number_format = '#,##0.00'


# ============================================================================
# SUMMARY GENERATOR
# ============================================================================

class SummaryGenerator:
    """Generates summary reports."""

    @staticmethod
    def generate_summary(
        reps: List[str],
        df_final: pd.DataFrame,
        df_called: pd.DataFrame
    ) -> pd.DataFrame:
        """Generate summary statistics for each sales rep."""
        summary_rows = []

        for rep in reps:
            # Visits data
            rep_visits = df_final[df_final["sales_rep_final"] == rep]
            customers_visited = rep_visits["customer_name_final"].nunique()
            order_value_visits = rep_visits["order_value"].sum()

            # Calls data
            rep_calls = df_called[df_called["sales_rep"] == rep]
            customers_called = rep_calls["customer_called"].nunique()
            order_value_calls = rep_calls["order_value"].sum()

            summary_rows.append({
                "SALESPERSON": rep,
                "CUSTOMERS VISITED": customers_visited,
                "ORDER VALUE FROM VISITS": order_value_visits,
                "CUSTOMERS CALLED": customers_called,
                "ORDER VALUE FROM CALLS": order_value_calls
            })

        return pd.DataFrame(summary_rows)

    @staticmethod
    def export_summary_text(df_summary: pd.DataFrame, filepath: str) -> None:
        """Export summary as formatted text for email."""
        summary_for_email = df_summary.copy()

        # Format numbers with commas
        summary_for_email["ORDER VALUE FROM VISITS"] = \
            summary_for_email["ORDER VALUE FROM VISITS"].map("{:,.2f}".format)
        summary_for_email["ORDER VALUE FROM CALLS"] = \
            summary_for_email["ORDER VALUE FROM CALLS"].map("{:,.2f}".format)

        with open(filepath, "w") as f:
            f.write(summary_for_email.to_string(index=False))

        logger.info(f"✅ Summary text saved: {filepath}")

    @staticmethod
    def export_summary_image(df_summary: pd.DataFrame, filepath: str) -> None:
        """Export summary as image (optional)."""
        if not HAS_DFI:
            logger.warning("dataframe_image not available, skipping image export")
            return

        try:
            dfi.export(df_summary, filepath)
            logger.info(f"✅ Summary image saved: {filepath}")
        except Exception as e:
            logger.error(f"Failed to export summary image: {e}")


# ============================================================================
# REP SHEET GENERATOR
# ============================================================================

class RepSheetGenerator:
    """Generates individual sales rep sheets."""

    def __init__(self, access_token: str, config: ReportConfig):
        self.access_token = access_token
        self.config = config

        # Initialize fill patterns
        self.customer_fill = PatternFill(
            start_color=config.customer_fill_color,
            end_color=config.customer_fill_color,
            fill_type="solid"
        )
        self.product_header_fill = PatternFill(
            start_color=config.product_header_color,
            end_color=config.product_header_color,
            fill_type="solid"
        )

    def create_rep_sheet(
        self,
        writer: pd.ExcelWriter,
        rep: str,
        df_final: pd.DataFrame,
        df_called: pd.DataFrame,
        orders_data: List[Dict]
    ) -> None:
        """Create a complete sheet for a sales rep."""
        logger.info(f"  Processing: {rep}")

        # Build customer dictionary
        rep_visits = df_final[df_final["sales_rep_final"] == rep]
        rep_calls = df_called[df_called["sales_rep"] == rep]
        customers = self._build_customer_dict(rep, rep_visits, rep_calls, orders_data)

        # Create sheet name
        sheet_name = rep.replace(".", "_").replace(" ", "_")[:31]

        # Create empty worksheet
        writer.book.create_sheet(sheet_name)
        ws = writer.book[sheet_name]

        # Write data with styling
        self._write_rep_data(ws, customers, rep)

        # Apply general styling (column widths and row heights)
        self._adjust_column_widths_and_heights(ws)

    def _build_customer_dict(
        self,
        rep: str,
        rep_visits: pd.DataFrame,
        rep_calls: pd.DataFrame,
        orders_data: List[Dict]
    ) -> Dict[str, Dict]:
        """Build dictionary of customer information."""
        customers = {}

        # Add visits
        for _, visit in rep_visits.iterrows():
            cust = visit["customer_name_final"]
            customers[cust] = {
                "visit_type": "Visited",
                "time_spent": visit.get("time_spent", ""),
                "orders": [],
            }

        # Add calls
        for _, call in rep_calls.iterrows():
            cust = call["customer_called"]
            customers.setdefault(
                cust,
                {"visit_type": "Called", "time_spent": "", "orders": []},
            )
            if customers[cust]["visit_type"] == "Visited":
                customers[cust]["visit_type"] = "Visited & Called"

        # Add orders
        for order in orders_data:
            if order.get("sales_rep") != rep:
                continue

            cust = order.get("customer_name")
            if not cust:
                continue

            customers.setdefault(
                cust,
                {"visit_type": "No Visit", "time_spent": "", "orders": []},
            )
            customers[cust]["orders"].append(order.get("id"))

        return customers

    def _write_rep_data(self, ws: Worksheet, customers: Dict[str, Dict], rep: str) -> None:
        """Write customer data to worksheet with styling."""
        row_idx = 1
        wrap_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Define border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for cust_name, info in customers.items():
            # Format customer name with visit type in brackets
            visit_type = info["visit_type"]
            if visit_type == "Visited":
                customer_display = f"{cust_name} (visited)"
            elif visit_type == "Called":
                customer_display = f"{cust_name} (called)"
            elif visit_type == "Visited & Called":
                customer_display = f"{cust_name} (visited & called)"
            else:
                customer_display = cust_name

            # Customer header row - write values first
            ws.cell(row=row_idx, column=1, value=customer_display)

            # Format time spent
            time_spent = info["time_spent"]
            if time_spent and visit_type in ["Visited", "Visited & Called"]:
                time_display = f"Time Spent: {time_spent}"
            else:
                time_display = ""

            ws.cell(row=row_idx, column=2, value=time_display)
            ws.cell(row=row_idx, column=3, value="")
            ws.cell(row=row_idx, column=4, value="")
            ws.cell(row=row_idx, column=5, value="")
            ws.cell(row=row_idx, column=6, value="")
            ws.cell(row=row_idx, column=7, value="")

            # Style customer header with Times New Roman font, wrap text, and borders
            for col_idx in range(1, 8):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = self.customer_fill
                cell.font = Font(
                    name=self.config.body_font_name,
                    size=self.config.body_font_size,
                    bold=True
                )
                cell.alignment = wrap_align
                cell.border = thin_border

            row_idx += 1

            # Fetch and add order items
            all_items = self._fetch_order_items(info["orders"], rep)

            if all_items:
                # Product table header row
                ws.cell(row=row_idx, column=1, value="Product ID")
                ws.cell(row=row_idx, column=2, value="Product Description")
                ws.cell(row=row_idx, column=3, value="")
                ws.cell(row=row_idx, column=4, value="Sold Qty")
                ws.cell(row=row_idx, column=5, value="Unit Cost")
                ws.cell(row=row_idx, column=6, value="Order Value")
                ws.cell(row=row_idx, column=7, value="")

                # Style product header with Times New Roman font, wrap text, and borders
                for col_idx in range(1, 8):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = self.product_header_fill
                    cell.font = Font(
                        name=self.config.body_font_name,
                        size=self.config.body_font_size,
                        bold=True,
                        color="FFFFFF"
                    )
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    cell.border = thin_border

                row_idx += 1

                # Add product rows with Times New Roman font, wrap text, and borders
                for item in all_items:
                    qty = float(item.get("sold_qty", 0))
                    cost = float(item.get("unit_cost", 0))

                    # Clean product description - remove "ID - " prefix
                    product_desc = item.get("product_desc", "")
                    product_id = str(item.get("product_id", ""))

                    # Remove product ID prefix if it exists
                    if product_desc and product_id:
                        prefix = f"{product_id} - "
                        if product_desc.startswith(prefix):
                            product_desc = product_desc[len(prefix):]

                    ws.cell(row=row_idx, column=1, value=product_id)
                    ws.cell(row=row_idx, column=2, value=product_desc)
                    ws.cell(row=row_idx, column=3, value="")
                    ws.cell(row=row_idx, column=4, value=qty)
                    ws.cell(row=row_idx, column=5, value=cost)
                    ws.cell(row=row_idx, column=6, value=f"{qty * cost:,.2f}")
                    ws.cell(row=row_idx, column=7, value="")

                    # Apply Times New Roman font, wrap text, and borders to product rows
                    for col_idx in range(1, 8):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font = Font(
                            name=self.config.body_font_name,
                            size=self.config.body_font_size
                        )
                        cell.alignment = wrap_align
                        cell.border = thin_border

                    row_idx += 1
            else:
                # No orders - set value FIRST, then merge
                no_orders_cell = ws.cell(row=row_idx, column=1)
                no_orders_cell.value = "No orders"
                no_orders_cell.font = Font(
                    name=self.config.body_font_name,
                    size=self.config.body_font_size,
                    color=self.config.error_color,
                    bold=True
                )
                no_orders_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                no_orders_cell.border = thin_border

                # Apply borders to all cells before merging
                for col_idx in range(1, 8):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border

                # Now merge cells
                ws.merge_cells(
                    start_row=row_idx,
                    start_column=1,
                    end_row=row_idx,
                    end_column=7
                )

                row_idx += 1

            # Empty separator row - add borders to maintain grid
            for col_idx in range(1, 8):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border

            row_idx += 1

    def _fetch_order_items(self, order_ids: List[int], rep: str) -> List[Dict]:
        """Fetch items for all order IDs."""
        all_items = []

        for order_id in order_ids:
            try:
                details = get_order_details(self.access_token, order_id)
                all_items.extend(details.get("entries", []))
            except Exception as e:
                logger.error(f"Error fetching order {order_id} for {rep}: {e}")
                # Don't add error rows - just log and continue

        return all_items

    def _adjust_column_widths_and_heights(self, ws: Worksheet) -> None:
        """Set wider column widths and calculate row heights for rep sheet."""
        column_widths = {
            1: 45,  # Column A: Customer Name with visit type (wider)
            2: 40,  # Column B: Time Spent / Product Description (wider for long descriptions)
            3: 20,  # Column C: Extra space
            4: 15,  # Column D: Sold Qty
            5: 15,  # Column E: Unit Cost
            6: 18,  # Column F: Order Value
            7: 10,  # Column G: Empty column
        }

        # Set column widths
        for col_idx, width in column_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = width

        # Calculate and set row heights for all rows
        styler = ExcelStyler(self.config)
        for row_idx in range(1, ws.max_row + 1):
            max_height = 18  # Default minimum height (increased from 15)

            for col_idx in range(1, 8):  # We have 7 columns
                cell = ws.cell(row=row_idx, column=col_idx)
                cell_value = cell.value

                # Skip empty cells
                if cell_value is None or (isinstance(cell_value, str) and not cell_value.strip()):
                    continue

                # Get column width for this column
                col_width = column_widths.get(col_idx, 20)

                # Calculate required height
                try:
                    height = styler.calculate_row_height(cell_value, col_width, self.config.body_font_size)
                    max_height = max(max_height, height)
                except Exception as e:
                    logger.warning(f"Could not calculate height for row {row_idx}, col {col_idx}: {e}")
                    max_height = max(max_height, 35)  # Fallback to larger height

            # Set the row height with extra padding for readability
            ws.row_dimensions[row_idx].height = max_height


# ============================================================================
# MAIN REPORT GENERATOR
# ============================================================================

def generate_detailed_report(
    visits_data: List[Dict],
    orders_data: List[Dict],
    config: ReportConfig = None
) -> None:
    """
    Generate detailed CSFA report with visits and orders.

    Args:
        visits_data: List of visit records
        orders_data: List of order records
        config: Optional configuration object
    """
    if config is None:
        config = ReportConfig.from_env()

    access_token = os.getenv("ACCESS_TOKEN")
    if not access_token:
        raise ValueError("ACCESS_TOKEN not found in environment variables")

    logger.info(f"🚀 Starting report generation: {config.output_file}")

    # Remove old file
    if os.path.exists(config.output_file):
        os.remove(config.output_file)
        logger.info(f"Removed old file: {config.output_file}")

    # Initialize processors
    processor = DataProcessor()
    styler = ExcelStyler(config)
    summary_gen = SummaryGenerator()
    rep_gen = RepSheetGenerator(access_token, config)

    # Process data
    logger.info("📊 Processing data...")
    df_visits = processor.clean_visits(visits_data)
    df_orders = processor.clean_orders(orders_data)
    df_final = processor.merge_visits_orders(df_visits, df_orders)
    df_called = processor.get_called_customers(df_visits, df_orders)
    reps = processor.get_sales_reps(df_final, df_called)

    logger.info(f"Found {len(reps)} sales representatives")

    # Generate summary
    df_summary = summary_gen.generate_summary(reps, df_final, df_called)

    # Create Excel file
    logger.info("📝 Creating Excel file...")
    with pd.ExcelWriter(config.output_file, engine="openpyxl") as writer:
        # Write summary sheet with special styling
        df_summary.to_excel(writer, index=False, sheet_name="Summary")
        ws = writer.sheets["Summary"]
        styler.apply_summary_styling(ws)  # Use summary-specific styling
        styler.format_money_columns(ws, [2, 4])  # Order value columns

        # Write individual rep sheets
        for rep in reps:
            rep_gen.create_rep_sheet(writer, rep, df_final, df_called, orders_data)

    # Export summary files
    summary_gen.export_summary_text(df_summary, config.summary_text_file)
    summary_gen.export_summary_image(df_summary, config.summary_image_file)

    logger.info(f"✅ Report generation complete: {config.output_file}")


# ============================================================================
# EXAMPLE USAGE
# ============================================================================

if __name__ == "__main__":
    # This allows testing the module independently
    logger.info("Report generator module loaded successfully")
